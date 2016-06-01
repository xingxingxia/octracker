import subprocess, xlwt, xlrd, sys
from openpyxl import load_workbook
class Handler(): 
	def __init__(self):
		pass
	def call(self, *args):
		command = ["oc"] + list(args) + ["-h"]
		outputStr = subprocess.check_output(command)
		return outputStr
	def get_command_title(self,outputStr):
		commandList = []
		for i in range(outputStr.count(':')):
			commandList.append(outputStr.split(':')[i].split('\n')[-1])
		return commandList
	def get_commands(self,pageStr,title):
		''' input: title - str, title name, e.g. "Basic Commands"
			output: (commandsL,descriptionL)  commandsL - list for commands of the title,  descriptionL - list for descriptions of commands
		'''
		commandsL = []
		descriptionL = []
		cmd_dspt_Str = pageStr.split(title+':')[1].split("\n\n")[0].strip()
		cmd_dscpt_Lists = cmd_dspt_Str.split("\n")
		for cmd_dscpt in cmd_dscpt_Lists:
			wordlist = cmd_dscpt.split()
			cmd = wordlist[0]
			wordlist.remove(wordlist[0])
			description = " ".join(wordlist)
			descriptionL.append(description)
			commandsL.append(cmd)
		return (commandsL, descriptionL)

class oc(Handler):
	def __init__(self):
		Handler.__init__(self)
		pass
	def call_command_title(self, *args):
		self.commandTitleStr = self.call(*args)
		self.commandTitleList = self.get_command_title(self.commandTitleStr)

	def open_xls(self,sheetName):
		#try:
		#	self.book = xlrd.open_workbook("octracker.xls")
		#	self.tablesheet = self.book.sheet_by_name(sheetName)
		#except:
		#	self.book = xlwt.Workbook()
		#	self.newsheet = self.book.add_sheet(sheetName)
		#	self.book.save("octracker.xls")
		#	self.book = xlrd.open_workbook("octracker.xls")
		#	self.tablesheet = self.book.sheet_by_name(sheetName)
		self.wb = load_workbook("octracker.xlsx")
		self.ws = wb.get_sheet_by_name(sheetName)
	

		self.start_row = 1

	def write_sheet(self,colValList,colWriteNum):
		for index in range(colValList.__len__()):

		#	row = self.tablesheet.row_values(self.start_row + index)     ----for xlwt
		#	row.write(colWriteNum, colValList[index])   ----for xlwt
		#	self.tablesheet.put_cell(row, colWriteNum, colValList[index])
		#	self.tablesheet.cell(row, colWriteNum).value == colValList[index]
			self.c = self.ws.cell(self.start_row + index, colWriteNum)
			self.c.value = colValList[index]

		self.start_row = self.start_row + index +1


if __name__ == '__main__':
	test=oc()
	test.call_command_title("")  #	whole command page - test.commandTitleStr
	test.open_xls("all")
	if len(sys.argv)<2: 
		colWriteNum=0
	else:
		colWriteNum=int(sys.argv[1])-1

	print "command titles are \n" + str(test.commandTitleList) + '\n====================================\n'
	for title in test.commandTitleList:
		cmd_dscpt = test.get_commands(pageStr=test.commandTitleStr, title=title)
		commandList = cmd_dscpt[0]
		print title + ':'
		print commandList
		print '\n'
		
		test.write_sheet(commandList, colWriteNum)
	wb.save("octracker.xlsx")



